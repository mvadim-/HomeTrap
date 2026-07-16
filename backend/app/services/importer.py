from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import UTC, date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import Apartment, Invoice, InvoiceLine, InvoiceStatus, Service, Tariff
from app.services.billing import (
    InvoiceChronologyError,
    serialize_invoice_mutations,
    validate_invoice_chronology,
)

MONEY = Decimal("0.01")
ZERO = Decimal("0")
ONE = Decimal("1")
MONTHS = {
    "січ": 1,
    "січень": 1,
    "лют": 2,
    "лютий": 2,
    "бер": 3,
    "березень": 3,
    "кві": 4,
    "квітень": 4,
    "тра": 5,
    "травень": 5,
    "чер": 6,
    "червень": 6,
    "лип": 7,
    "липень": 7,
    "сер": 8,
    "серпень": 8,
    "вер": 9,
    "вересень": 9,
    "жов": 10,
    "жовтень": 10,
    "лис": 11,
    "листопад": 11,
    "гру": 12,
    "грудень": 12,
}

LEGACY_SERVICE_ALIASES = {
    "доставка газу": "газ доставка",
    "газ доставка": "газ доставка",
    "водопостачання": "вода",
    "вода": "вода",
    "вода абонентська плата": "вода доставка",
    "вода доставка": "вода доставка",
    "інтернет": "інтернет (uteam)",
    "інтернет (uteam)": "інтернет (uteam)",
}
LEGACY_UNITS = {
    "газ": "м³",
    "світло": "кВт·год",
    "вода": "м³",
}


class ImportFormatError(ValueError):
    pass


@dataclass
class ImportReport:
    invoices_created: int = 0
    invoices_skipped: int = 0
    services_created: int = 0
    tariffs_created: int = 0
    warnings: list[str] = field(default_factory=list)


def _normalized(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip()).casefold()


def parse_decimal(value: object) -> Decimal | None:
    if value is None or _normalized(value) in {"", "-", "—", "–"}:
        return None
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    text = str(value).replace("\xa0", "").replace(" ", "").strip().replace(",", ".")
    match = re.search(r"[-+]?\d+(?:\.\d+)?", text)
    if match is None:
        raise InvalidOperation
    return Decimal(match.group(0))


def _period(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date().replace(day=1)
    if isinstance(value, date):
        return value.replace(day=1)
    text = _normalized(value).replace(".", " ")
    match = re.search(r"([а-яіїє]+)\s+(20\d{2})", text)
    if not match:
        return None
    month = MONTHS.get(match.group(1))
    return date(int(match.group(2)), month, 1) if month else None


def _header_map(row: tuple[object, ...]) -> dict[str, int]:
    aliases = {
        "service": ("послуга", "назва послуги", "назва"),
        "kind": ("тип", "вид"),
        "unit": ("одиниця", "од. вим.", "одиниця виміру"),
        "account": ("особовий рахунок", "рахунок постачальника"),
        "previous": ("попередній", "попередній показник"),
        "current": ("поточний", "поточний показник"),
        "tariff": ("тариф",),
        "amount": ("сума", "до сплати"),
    }
    result: dict[str, int] = {}
    for index, value in enumerate(row):
        name = _normalized(value)
        for key, variants in aliases.items():
            if name in variants:
                result[key] = index
    return result


def _rows(sheet: Worksheet) -> list[tuple[object, ...]]:
    return [tuple(cell.value for cell in row) for row in sheet.iter_rows()]


def _find_table(sheet: Worksheet, required: set[str]) -> tuple[int, dict[str, int], list[tuple[object, ...]]]:
    rows = _rows(sheet)
    for index, row in enumerate(rows):
        mapping = _header_map(row)
        if required <= mapping.keys():
            return index, mapping, rows
    raise ImportFormatError(f"Не знайдено таблицю у вкладці «{sheet.title}»")


def _value(row: tuple[object, ...], mapping: dict[str, int], key: str) -> object:
    index = mapping.get(key)
    return row[index] if index is not None and index < len(row) else None


def _decimal_or_warning(
    value: object,
    *,
    sheet: str,
    field_name: str,
    report: ImportReport,
) -> Decimal | None:
    try:
        return parse_decimal(value)
    except InvalidOperation:
        report.warnings.append(
            f"{sheet}: нечислове значення «{value}» у полі «{field_name}»",
        )
        return None


def _tariff_or_warning(
    value: object,
    *,
    sheet: str,
    field_name: str,
    report: ImportReport,
    fallback: Decimal | None = None,
) -> Decimal | None:
    tariff = _decimal_or_warning(
        value,
        sheet=sheet,
        field_name=field_name,
        report=report,
    )
    tariff = fallback if tariff is None else tariff
    if tariff is not None and tariff <= 0:
        raise ImportFormatError(f"{sheet}: {field_name} має бути додатним")
    return tariff


def _find_information_sheet(sheet_names: list[str]) -> str:
    for name in sheet_names:
        if _normalized(name) in {"загальна інформація", "загальна iнформацiя"}:
            return name
    raise ImportFormatError("Відсутня вкладка «Загальна інформація»")


def _legacy_service_key(value: object) -> str:
    normalized = _normalized(value)
    return LEGACY_SERVICE_ALIASES.get(normalized, normalized)


def _is_legacy_export(sheet: Worksheet) -> bool:
    return _normalized(sheet.cell(1, 1).value) == "тарифи на комунальні послуги"


def _legacy_month_rows(sheet: Worksheet) -> list[tuple[object, ...]]:
    rows = _rows(sheet)
    if not rows or _normalized(rows[0][0]) != "комунальні послуги":
        raise ImportFormatError(f"Невідомий формат вкладки «{sheet.title}»")
    result: list[tuple[object, ...]] = []
    for row in rows[2:]:
        name = _normalized(row[0] if row else None)
        if name == "разом":
            return result
        if name:
            result.append(row)
    raise ImportFormatError(f"Не знайдено підсумок у вкладці «{sheet.title}»")


def _legacy_appearances(
    workbook,
    information_name: str,
) -> dict[str, tuple[str, bool]]:
    result: dict[str, tuple[str, bool]] = {}
    for name in workbook.sheetnames:
        if name == information_name or _period(name) is None:
            continue
        for row in _legacy_month_rows(workbook[name]):
            display_name = str(row[0]).strip()
            key = _legacy_service_key(display_name)
            try:
                previous = parse_decimal(row[1] if len(row) > 1 else None)
                current = parse_decimal(row[2] if len(row) > 2 else None)
            except InvalidOperation:
                previous = current = None
            metered = previous is not None or current is not None
            existing = result.get(key)
            result[key] = (display_name, metered or (existing[1] if existing else False))
    return result


def _cell_text(value: object) -> str | None:
    if value is None or _normalized(value) in {"", "-", "—", "–"}:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip() or None


def _import_legacy_services(
    session: Session,
    apartment: Apartment,
    workbook,
    information_name: str,
    report: ImportReport,
) -> dict[str, Service]:
    sheet = workbook[information_name]
    rows = _rows(sheet)
    if len(rows) < 3:
        raise ImportFormatError("Вкладка «Загальна інформація» порожня")
    appearances = _legacy_appearances(workbook, information_name)
    known: dict[str, Service] = {}
    for service in apartment.services:
        known[_legacy_service_key(service.name)] = service
        known[_normalized(service.name)] = service
    tariff_columns = [
        (index, period)
        for index, value in enumerate(rows[1])
        if (period := _period(value)) is not None
    ]
    if not tariff_columns:
        raise ImportFormatError("Не знайдено дат тарифів у вкладці «Загальна інформація»")

    for row in rows[2:]:
        raw_name = row[0] if row else None
        if _normalized(raw_name) == "курс валют":
            break
        if raw_name is None or not _normalized(raw_name):
            continue
        info_name = str(raw_name).strip()
        key = _legacy_service_key(info_name)
        appearance = appearances.get(key)
        display_name = appearance[0] if appearance else info_name
        service = known.get(key)
        if service is None:
            service = Service(
                apartment=apartment,
                name=display_name,
                kind="metered" if appearance and appearance[1] else "fixed",
                unit=LEGACY_UNITS.get(key),
                provider_account=_cell_text(row[1] if len(row) > 1 else None),
                sort_order=max((item.sort_order for item in known.values()), default=-1) + 1,
                is_active=appearance is not None,
            )
            session.add(service)
            session.flush()
            report.services_created += 1
        known[key] = service
        known[_normalized(info_name)] = service
        known[_normalized(display_name)] = service
        existing_dates = {tariff.valid_from for tariff in service.tariffs}
        for column, valid_from in tariff_columns:
            raw = row[column] if column < len(row) else None
            tariff_value = _tariff_or_warning(
                raw,
                sheet=sheet.title,
                field_name=f"тариф для «{display_name}» за {valid_from:%Y-%m}",
                report=report,
            )
            if tariff_value is None or valid_from in existing_dates:
                continue
            session.add(Tariff(service=service, value=tariff_value, valid_from=valid_from))
            existing_dates.add(valid_from)
            report.tariffs_created += 1
    return known


def _import_services(
    session: Session,
    apartment: Apartment,
    sheet: Worksheet,
    report: ImportReport,
) -> dict[str, Service]:
    header_index, mapping, rows = _find_table(sheet, {"service", "kind"})
    known = {_normalized(service.name): service for service in apartment.services}
    header = rows[header_index]
    tariff_columns = [
        (index, period)
        for index, value in enumerate(header)
        if (period := _period(value)) is not None
    ]
    for row in rows[header_index + 1 :]:
        name_value = _value(row, mapping, "service")
        name = str(name_value).strip() if name_value is not None else ""
        if not name:
            continue
        service = known.get(_normalized(name))
        if service is None:
            kind_value = _normalized(_value(row, mapping, "kind"))
            kind_aliases = {
                "metered": "metered",
                "лічильник": "metered",
                "за лічильником": "metered",
                "fixed": "fixed",
                "фіксована": "fixed",
                "фіксований": "fixed",
            }
            kind = kind_aliases.get(kind_value)
            if kind is None:
                raise ImportFormatError(
                    f"{sheet.title}: невідомий тип послуги «{_value(row, mapping, 'kind')}» для «{name}»"
                )
            service = Service(
                apartment=apartment,
                name=name,
                kind=kind,
                unit=str(_value(row, mapping, "unit")).strip() or None
                if _value(row, mapping, "unit") is not None
                else None,
                provider_account=str(_value(row, mapping, "account")).strip() or None
                if _value(row, mapping, "account") is not None
                else None,
                sort_order=len(known),
            )
            session.add(service)
            session.flush()
            known[_normalized(name)] = service
            report.services_created += 1
        existing_dates = {tariff.valid_from for tariff in service.tariffs}
        for column, valid_from in tariff_columns:
            raw = row[column] if column < len(row) else None
            tariff_value = _tariff_or_warning(
                raw,
                sheet=sheet.title,
                field_name=f"тариф для «{name}» за {valid_from:%Y-%m}",
                report=report,
            )
            if tariff_value is None or valid_from in existing_dates:
                continue
            session.add(Tariff(service=service, value=tariff_value, valid_from=valid_from))
            existing_dates.add(valid_from)
            report.tariffs_created += 1
    return known


def _metadata(sheet: Worksheet) -> dict[str, object]:
    result: dict[str, object] = {}
    aliases = {
        "rate": ("курс", "курс нбу", "курс usd"),
        "rent_usd": ("оренда usd", "оренда, usd", "оренда $"),
        "rent_uah": ("оренда грн", "оренда, грн"),
    }
    for row in _rows(sheet):
        for index, cell in enumerate(row):
            name = _normalized(cell)
            for key, variants in aliases.items():
                if name in variants and index + 1 < len(row):
                    result[key] = row[index + 1]
    return result


def _active_tariff(service: Service, period: date) -> Decimal | None:
    candidates = [tariff for tariff in service.tariffs if tariff.valid_from <= period]
    return max(candidates, key=lambda item: item.valid_from).value if candidates else None


def _legacy_month_metadata(
    sheet: Worksheet,
) -> tuple[dict[str, object], object | None, str | None]:
    rows = _rows(sheet)
    rent_heading_index = next(
        (index for index, row in enumerate(rows) if _normalized(row[0]) == "оренда"),
        None,
    )
    final_index = next(
        (index for index, row in enumerate(rows) if _normalized(row[0]) == "разом до оплати"),
        None,
    )
    if rent_heading_index is None or final_index is None:
        raise ImportFormatError(f"Не знайдено секцію оренди у вкладці «{sheet.title}»")
    rent_index = next(
        (
            index
            for index in range(rent_heading_index + 1, final_index)
            if len(rows[index]) > 5
            and any(rows[index][column] not in (None, "") for column in (1, 2, 5))
        ),
        None,
    )
    if rent_index is None:
        raise ImportFormatError(f"Не знайдено дані оренди у вкладці «{sheet.title}»")
    adjustment_names = [
        str(rows[index][0]).strip()
        for index in range(rent_index + 1, final_index)
        if rows[index][0] not in (None, "") and len(rows[index]) > 5 and rows[index][5] not in (None, "")
    ]
    rent_row = rows[rent_index]
    final_row = rows[final_index]
    return (
        {
            "rent_usd": rent_row[1] if len(rent_row) > 1 else None,
            "rate": rent_row[2] if len(rent_row) > 2 else None,
            "rent_uah": rent_row[5] if len(rent_row) > 5 else None,
        },
        final_row[5] if len(final_row) > 5 else None,
        " / ".join(adjustment_names) or None,
    )


def _legacy_adjustment_service(
    session: Session,
    apartment: Apartment,
    services: dict[str, Service],
    name: str,
    report: ImportReport,
) -> Service:
    key = _legacy_service_key(name)
    service = services.get(key)
    if service is not None:
        return service
    service = Service(
        apartment=apartment,
        name=name,
        kind="fixed",
        sort_order=max((item.sort_order for item in services.values()), default=-1) + 1,
        is_active=False,
    )
    session.add(service)
    services[key] = service
    services[_normalized(name)] = service
    report.services_created += 1
    return service


def _record_legacy_tariff(
    session: Session,
    service: Service,
    period: date,
    value: Decimal,
    report: ImportReport,
) -> None:
    if not service.is_active or any(tariff.valid_from == period for tariff in service.tariffs):
        return
    if _active_tariff(service, period) == value:
        return
    session.add(Tariff(service=service, value=value, valid_from=period))
    report.tariffs_created += 1


def _import_month(
    session: Session,
    apartment: Apartment,
    services: dict[str, Service],
    sheet: Worksheet,
    period: date,
    report: ImportReport,
    *,
    legacy: bool = False,
) -> None:
    existing = session.scalar(
        select(Invoice).where(
            Invoice.apartment_id == apartment.id,
            Invoice.period == period,
        ),
    )
    if existing is not None:
        report.invoices_skipped += 1
        return
    try:
        validate_invoice_chronology(session, apartment.id, period)
    except InvoiceChronologyError as error:
        messages = {
            "later_invoice": "не можна імпортувати місяць перед наявним пізнішим рахунком",
            "earlier_draft": "не можна імпортувати місяць після незавершеної ранньої чернетки",
        }
        raise ImportFormatError(f"{sheet.title}: {messages[error.code]}") from error

    if legacy:
        mapping = {"service": 0, "previous": 1, "current": 2, "tariff": 4, "amount": 5}
        month_rows = _legacy_month_rows(sheet)
        metadata, legacy_final_total, legacy_adjustment_name = _legacy_month_metadata(sheet)
    else:
        header_index, mapping, rows = _find_table(sheet, {"service", "amount"})
        month_rows = rows[header_index + 1 :]
        metadata = _metadata(sheet)
        legacy_final_total = None
        legacy_adjustment_name = None
    rate = _decimal_or_warning(
        metadata.get("rate"), sheet=sheet.title, field_name="курс", report=report
    )
    rent_usd = _decimal_or_warning(
        metadata.get("rent_usd"), sheet=sheet.title, field_name="оренда USD", report=report
    )
    if rent_usd is not None and rent_usd < 0:
        raise ImportFormatError(f"{sheet.title}: оренда USD не може бути від’ємною")
    if rent_usd is None:
        rent_usd = apartment.rent_amount
    rent_uah = _decimal_or_warning(
        metadata.get("rent_uah"), sheet=sheet.title, field_name="оренда грн", report=report
    )
    if rent_uah is not None and rent_uah < 0:
        raise ImportFormatError(f"{sheet.title}: оренда грн не може бути від’ємною")
    if rent_uah is None:
        if rate is None or rate <= 0:
            raise ImportFormatError(
                f"{sheet.title}: потрібен додатний курс, якщо сума оренди у гривнях відсутня"
            )
        rent_uah = (rent_usd * rate).quantize(MONEY, rounding=ROUND_HALF_UP)
    resolved_rate = rate if rate is not None and rate > 0 else ZERO

    invoice = Invoice(
        apartment=apartment,
        period=period,
        status=InvoiceStatus.PAID.value,
        issued_at=datetime.now(UTC),
        paid_at=datetime.now(UTC),
        exchange_rate=resolved_rate,
        rent_amount_usd=rent_usd.quantize(MONEY, rounding=ROUND_HALF_UP),
        rent_amount_uah=rent_uah.quantize(MONEY, rounding=ROUND_HALF_UP),
        utilities_total=ZERO,
        grand_total=ZERO,
    )
    report.invoices_created += 1
    utilities_total = ZERO
    for row in month_rows:
        name_value = _value(row, mapping, "service")
        name = str(name_value).strip() if name_value is not None else ""
        if not name:
            continue
        service = services.get(_legacy_service_key(name) if legacy else _normalized(name))
        adjustment = False
        if service is None:
            if not legacy:
                raise ImportFormatError(f"{sheet.title}: невідома послуга «{name}»")
            service = _legacy_adjustment_service(session, apartment, services, name, report)
            adjustment = True
        previous = _decimal_or_warning(
            _value(row, mapping, "previous"),
            sheet=sheet.title,
            field_name=f"{name}: попередній показник",
            report=report,
        )
        current = _decimal_or_warning(
            _value(row, mapping, "current"),
            sheet=sheet.title,
            field_name=f"{name}: поточний показник",
            report=report,
        )
        tariff = ONE if adjustment else _tariff_or_warning(
            _value(row, mapping, "tariff"),
            sheet=sheet.title,
            field_name=f"{name}: тариф",
            report=report,
            fallback=_active_tariff(service, period),
        )
        if tariff is None:
            raise ImportFormatError(f"{sheet.title}: {name}: потрібен додатний тариф")
        if legacy and not adjustment:
            _record_legacy_tariff(session, service, period, tariff, report)
        amount = _decimal_or_warning(
            _value(row, mapping, "amount"),
            sheet=sheet.title,
            field_name=f"{name}: сума",
            report=report,
        )
        consumed = current - previous if current is not None and previous is not None else None
        if amount is None:
            amount = consumed * tariff if consumed is not None else ZERO
        amount = amount.quantize(MONEY, rounding=ROUND_HALF_UP)
        utilities_total += amount
        invoice.lines.append(
            InvoiceLine(
                service=service,
                service_name=service.name,
                service_kind=service.kind,
                prev_reading=previous,
                curr_reading=current,
                consumed=consumed,
                tariff_value=tariff,
                amount=amount,
            )
        )
    if legacy and legacy_final_total is not None:
        target_total = _decimal_or_warning(
            legacy_final_total,
            sheet=sheet.title,
            field_name="разом до оплати",
            report=report,
        )
        if target_total is not None:
            target_total = target_total.quantize(MONEY, rounding=ROUND_HALF_UP)
            adjustment_amount = target_total - invoice.rent_amount_uah - utilities_total
            if legacy_adjustment_name and adjustment_amount != ZERO:
                adjustment_service = _legacy_adjustment_service(
                    session,
                    apartment,
                    services,
                    legacy_adjustment_name,
                    report,
                )
                invoice.lines.append(
                    InvoiceLine(
                        service=adjustment_service,
                        service_name=adjustment_service.name,
                        service_kind="fixed",
                        prev_reading=None,
                        curr_reading=None,
                        consumed=None,
                        tariff_value=ONE,
                        amount=adjustment_amount,
                    )
                )
                utilities_total += adjustment_amount
    invoice.utilities_total = utilities_total.quantize(MONEY, rounding=ROUND_HALF_UP)
    invoice.grand_total = (invoice.rent_amount_uah + invoice.utilities_total).quantize(
        MONEY, rounding=ROUND_HALF_UP
    )
    session.add(invoice)
    session.flush()


def _warn_month_gaps(periods: list[date], report: ImportReport) -> None:
    for previous, current in zip(periods, periods[1:], strict=False):
        next_month = date(previous.year + (previous.month == 12), previous.month % 12 + 1, 1)
        if current != next_month:
            report.warnings.append(
                f"Розрив між місяцями {previous:%Y-%m} та {current:%Y-%m}",
            )


def import_xlsx(
    session: Session,
    apartment: Apartment,
    content: bytes,
    *,
    dry_run: bool = False,
) -> ImportReport:
    try:
        workbook = load_workbook(BytesIO(content), data_only=True)
    except Exception as error:
        raise ImportFormatError("Не вдалося прочитати XLSX-файл") from error
    report = ImportReport()
    information_name = _find_information_sheet(workbook.sheetnames)
    legacy = _is_legacy_export(workbook[information_name])
    serialize_invoice_mutations(session, apartment.id)
    services = (
        _import_legacy_services(session, apartment, workbook, information_name, report)
        if legacy
        else _import_services(session, apartment, workbook[information_name], report)
    )
    months = sorted(
        (period, workbook[name])
        for name in workbook.sheetnames
        if name != information_name and (period := _period(name)) is not None
    )
    if not months:
        raise ImportFormatError("Не знайдено вкладок у форматі «<Місяць> <Рік>»")
    _warn_month_gaps([period for period, _ in months], report)
    for period, sheet in months:
        _import_month(session, apartment, services, sheet, period, report, legacy=legacy)
    if dry_run:
        session.rollback()
    else:
        session.commit()
    return report
