from __future__ import annotations

from concurrent.futures import ThreadPoolExecutor
from datetime import date
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from threading import Event

from httpx import ASGITransport, AsyncClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import func, select

from app.config import Settings
from app.db import create_database_engine, create_session_factory
from app.main import create_app
from app.models import Apartment, Invoice, InvoiceLine, Service, Tariff
import app.services.billing as billing_service
import app.services.importer as importer_service
from app.services.billing import create_draft
from app.services.importer import ImportFormatError, import_xlsx

FIXTURE = Path(__file__).parent / "fixtures" / "sample_import.xlsx"


def _legacy_export_content() -> bytes:
    workbook = Workbook()
    information = workbook.active
    information.title = "Загальна інформація"
    information["A1"] = "Тарифи на комунальні послуги"
    information["B2"] = "Номер рахунку"
    information["D2"] = "Квітень 2024"
    information["E2"] = "Травень 2024"
    information.append(["Газ", "ACC-GAS", None, 7.95689, 7.95689])
    information.append(["Водопостачання", "ACC-WATER", None, 28.24, 28.24])
    information.append(["Інтернет", "ACC-NET", None, 240, 240])
    information["A8"] = "Курс Валют"

    for title, previous, current, water_previous, water_current, rate, adjustment in (
        ("Квітень 2024", 100, 122, 10, 15, 44.68, False),
        ("Травень 2024", 122, 132, 15, 19, 45, True),
    ):
        sheet = workbook.create_sheet(title)
        sheet.append(["Комунальні послуги", "Показники", None, None, "Тариф", "До оплати"])
        sheet.append([None, "Попередні", "Поточні", "Спожито"])
        gas_amount = (current - previous) * 7.95689
        water_amount = (water_current - water_previous) * 28.24
        sheet.append(["Газ", previous, current, current - previous, 7.95689, gas_amount])
        sheet.append([
            "Вода",
            water_previous,
            water_current,
            water_current - water_previous,
            28.24,
            water_amount,
        ])
        sheet.append(["Інтернет (uteam)", "-", "-", None, 240, 240])
        utility_total = Decimal(str(gas_amount)) + Decimal(str(water_amount)) + Decimal("240")
        if adjustment:
            sheet.append(["Виклик майстра", None, None, None, None, -50])
            utility_total -= Decimal("50")
        sheet.append(["Разом", None, None, None, None, float(utility_total)])
        sheet.append([])
        sheet.append(["Оренда", "Вартість", "Курс"])
        rent_uah = Decimal("325") * Decimal(str(rate))
        sheet.append([None, 325, rate, None, None, float(rent_uah)])
        expected_total = utility_total + rent_uah
        if adjustment:
            sheet.append(["Завдаток", None, None, None, None, 500])
            expected_total -= Decimal("500")
        sheet.append(["Разом до оплати", None, None, None, None, float(expected_total)])

    content = BytesIO()
    workbook.save(content)
    return content.getvalue()


async def _client(tmp_path):
    settings = Settings(
        database_path=tmp_path / "import.db",
        secret_key="test-session-secret",
        debug=True,
        scheduler_enabled=False,
        admin_username="admin",
        admin_password="password",
    )
    application = create_app(settings)
    lifespan = application.router.lifespan_context(application)
    await lifespan.__aenter__()
    client = AsyncClient(transport=ASGITransport(app=application), base_url="http://test")
    login = await client.post(
        "/api/auth/login",
        json={"username": "admin", "password": "password"},
    )
    assert login.status_code == 200
    return application, lifespan, client


async def _close(lifespan, client: AsyncClient) -> None:
    await client.aclose()
    await lifespan.__aexit__(None, None, None)


def _apartment(application) -> int:
    engine = create_database_engine(application.state.settings.database_path)
    with create_session_factory(engine)() as session:
        apartment = Apartment(
            name="Тестова квартира",
            address="Анонімізовано",
            rent_amount=325,
            rent_currency="USD",
        )
        session.add(apartment)
        session.commit()
        result = apartment.id
    engine.dispose()
    return result


async def _upload(client: AsyncClient, apartment_id: int, content: bytes, **params):
    return await client.post(
        f"/api/apartments/{apartment_id}/import",
        params=params,
        files={
            "file": (
                "history.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )


async def test_full_import_handles_export_artifacts_and_broken_cells(tmp_path) -> None:
    application, lifespan, client = await _client(tmp_path)
    try:
        apartment_id = _apartment(application)
        response = await _upload(client, apartment_id, FIXTURE.read_bytes())
        assert response.status_code == 200
        report = response.json()
        assert report["invoices_created"] == 2
        assert report["services_created"] == 2
        assert report["tariffs_created"] == 4
        assert any("нечислове значення «не число»" in item for item in report["warnings"])

        engine = create_database_engine(application.state.settings.database_path)
        with create_session_factory(engine)() as session:
            invoices = list(session.scalars(select(Invoice).order_by(Invoice.period)))
            assert [invoice.status for invoice in invoices] == ["paid", "paid"]
            assert str(invoices[0].exchange_rate) == "44.680000"
            assert str(invoices[0].rent_amount_uah) == "14521.00"
            assert str(invoices[0].utilities_total) == "385.56"
            assert str(invoices[0].grand_total) == "14906.56"
            assert str(invoices[0].lines[0].tariff_value) == "7.95689"
            assert invoices[0].lines[1].prev_reading is None
            assert session.scalar(select(func.count(Service.id))) == 2
            assert session.scalar(select(func.count(Tariff.id))) == 4
        engine.dispose()
    finally:
        await _close(lifespan, client)


async def test_import_supports_real_export_layout_and_dry_run(tmp_path) -> None:
    application, lifespan, client = await _client(tmp_path)
    try:
        apartment_id = _apartment(application)
        content = _legacy_export_content()

        preview = await _upload(client, apartment_id, content, dry_run="true")
        assert preview.status_code == 200
        assert preview.json()["invoices_created"] == 2

        engine = create_database_engine(application.state.settings.database_path)
        with create_session_factory(engine)() as session:
            assert session.scalar(select(func.count(Invoice.id))) == 0
            assert session.scalar(select(func.count(Service.id))) == 0
        engine.dispose()

        imported = await _upload(client, apartment_id, content)
        assert imported.status_code == 200
        assert imported.json()["invoices_created"] == 2

        engine = create_database_engine(application.state.settings.database_path)
        with create_session_factory(engine)() as session:
            invoices = list(session.scalars(select(Invoice).order_by(Invoice.period)))
            assert str(invoices[1].grand_total) == "14507.53"
            lines = list(
                session.scalars(
                    select(InvoiceLine)
                    .where(InvoiceLine.invoice_id == invoices[1].id)
                    .order_by(InvoiceLine.id)
                )
            )
            amounts = {line.service_name: str(line.amount) for line in lines}
            assert amounts["Виклик майстра"] == "-50.00"
            assert amounts["Завдаток"] == "-500.00"
            inactive = list(session.scalars(select(Service).where(Service.is_active.is_(False))))
            assert {service.name for service in inactive} == {"Виклик майстра", "Завдаток"}
        engine.dispose()
    finally:
        await _close(lifespan, client)


async def test_dry_run_previews_without_writing_and_import_is_idempotent(tmp_path) -> None:
    application, lifespan, client = await _client(tmp_path)
    try:
        apartment_id = _apartment(application)
        content = FIXTURE.read_bytes()
        preview = await _upload(client, apartment_id, content, dry_run="true")
        assert preview.status_code == 200
        assert preview.json()["invoices_created"] == 2

        engine = create_database_engine(application.state.settings.database_path)
        with create_session_factory(engine)() as session:
            assert session.scalar(select(func.count(Invoice.id))) == 0
            assert session.scalar(select(func.count(Service.id))) == 0
        engine.dispose()

        imported = await _upload(client, apartment_id, content)
        engine = create_database_engine(application.state.settings.database_path)
        with create_session_factory(engine)() as session:
            first_invoice = session.scalar(select(Invoice).order_by(Invoice.id))
            first_invoice.status = "issued"
            first_invoice.rent_amount_uah = 1
            first_line = session.scalar(select(InvoiceLine).order_by(InvoiceLine.id))
            session.delete(first_line)
            session.commit()
        engine.dispose()
        repeated = await _upload(client, apartment_id, content)
        assert imported.json()["invoices_created"] == 2
        assert repeated.json()["invoices_created"] == 0
        assert repeated.json()["invoices_skipped"] == 2
        assert repeated.json()["services_created"] == 0
        assert repeated.json()["tariffs_created"] == 0
        engine = create_database_engine(application.state.settings.database_path)
        with create_session_factory(engine)() as session:
            preserved = session.scalar(select(Invoice).order_by(Invoice.id))
            assert preserved.status == "issued"
            assert str(preserved.rent_amount_uah) == "1.00"
            assert session.scalar(select(func.count(InvoiceLine.id))) == 3
        engine.dispose()
    finally:
        await _close(lifespan, client)


async def test_gap_warning_merged_cells_and_upload_errors(tmp_path) -> None:
    workbook = load_workbook(FIXTURE)
    assert "A1:F1" in {str(item) for item in workbook["Загальна інформація"].merged_cells.ranges}
    workbook["Тра 2024"].title = "Чер 2024"
    content = BytesIO()
    workbook.save(content)

    application, lifespan, client = await _client(tmp_path)
    try:
        apartment_id = _apartment(application)
        response = await _upload(client, apartment_id, content.getvalue())
        assert response.status_code == 200
        assert any("Розрив між місяцями" in item for item in response.json()["warnings"])

        missing = await _upload(client, apartment_id + 999, FIXTURE.read_bytes())
        assert missing.status_code == 404
        wrong_extension = await client.post(
            f"/api/apartments/{apartment_id}/import",
            files={"file": ("history.txt", b"broken", "text/plain")},
        )
        assert wrong_extension.status_code == 422
        broken = await _upload(client, apartment_id, b"not an xlsx")
        assert broken.status_code == 422
    finally:
        await _close(lifespan, client)


async def test_import_requires_authentication(tmp_path) -> None:
    application, lifespan, client = await _client(tmp_path)
    try:
        apartment_id = _apartment(application)
        await client.post("/api/auth/logout")
        response = await _upload(client, apartment_id, FIXTURE.read_bytes())
        assert response.status_code == 401
    finally:
        await _close(lifespan, client)


async def test_import_rejects_unknown_service_kind_and_missing_required_rate(tmp_path) -> None:
    application, lifespan, client = await _client(tmp_path)
    try:
        apartment_id = _apartment(application)
        workbook = load_workbook(FIXTURE)
        workbook["Загальна інформація"]["B4"] = "невідомий"
        content = BytesIO()
        workbook.save(content)
        response = await _upload(client, apartment_id, content.getvalue())
        assert response.status_code == 422
        assert "невідомий тип послуги" in response.json()["detail"]

        workbook = load_workbook(FIXTURE)
        month = workbook["Кві 2024"]
        month["B3"] = None
        month["B5"] = None
        content = BytesIO()
        workbook.save(content)
        response = await _upload(client, apartment_id, content.getvalue())
        assert response.status_code == 422
        assert "потрібен додатний курс" in response.json()["detail"]
    finally:
        await _close(lifespan, client)


async def test_import_rejects_non_positive_tariffs_without_writing(tmp_path) -> None:
    application, lifespan, client = await _client(tmp_path)
    try:
        apartment_id = _apartment(application)
        for invalid_tariff in (0, -7.5):
            workbook = load_workbook(FIXTURE)
            workbook["Загальна інформація"]["E4"] = invalid_tariff
            content = BytesIO()
            workbook.save(content)

            response = await _upload(client, apartment_id, content.getvalue())

            assert response.status_code == 422
            assert "тариф для «Газ»" in response.json()["detail"]
            assert "має бути додатним" in response.json()["detail"]

            engine = create_database_engine(
                application.state.settings.database_path
            )
            with create_session_factory(engine)() as session:
                assert session.scalar(select(func.count(Service.id))) == 0
                assert session.scalar(select(func.count(Tariff.id))) == 0
            engine.dispose()
    finally:
        await _close(lifespan, client)


async def test_import_rejects_non_positive_month_tariff_without_fallback(tmp_path) -> None:
    application, lifespan, client = await _client(tmp_path)
    try:
        apartment_id = _apartment(application)
        for invalid_tariff in (0, -7.5):
            workbook = load_workbook(FIXTURE)
            workbook["Кві 2024"]["D8"] = invalid_tariff
            content = BytesIO()
            workbook.save(content)

            response = await _upload(client, apartment_id, content.getvalue())

            assert response.status_code == 422
            assert "Газ: тариф має бути додатним" in response.json()["detail"]

            engine = create_database_engine(
                application.state.settings.database_path
            )
            with create_session_factory(engine)() as session:
                assert session.scalar(select(func.count(Service.id))) == 0
                assert session.scalar(select(func.count(Invoice.id))) == 0
            engine.dispose()
    finally:
        await _close(lifespan, client)


async def test_import_rejects_negative_rent_without_writing(tmp_path) -> None:
    application, lifespan, client = await _client(tmp_path)
    try:
        apartment_id = _apartment(application)
        for cell, field_name in (("B4", "оренда USD"), ("B5", "оренда грн")):
            workbook = load_workbook(FIXTURE)
            workbook["Кві 2024"][cell] = -1
            content = BytesIO()
            workbook.save(content)

            response = await _upload(client, apartment_id, content.getvalue())

            assert response.status_code == 422
            assert f"{field_name} не може бути від’ємною" in response.json()["detail"]
            engine = create_database_engine(application.state.settings.database_path)
            with create_session_factory(engine)() as session:
                assert session.scalar(select(func.count(Invoice.id))) == 0
                assert session.scalar(select(func.count(Service.id))) == 0
            engine.dispose()
    finally:
        await _close(lifespan, client)


def test_concurrent_draft_serializes_before_xlsx_import(
    db_engine,
    monkeypatch,
) -> None:
    session_factory = create_session_factory(db_engine)
    with session_factory() as session:
        apartment = Apartment(
            name="Тестова квартира",
            address="Анонімізовано",
            rent_amount=Decimal("325.00"),
            rent_currency="USD",
        )
        session.add(apartment)
        session.commit()
        apartment_id = apartment.id

    draft_validated = Event()
    import_attempted = Event()
    import_validated = Event()
    original_validate = billing_service.validate_invoice_chronology

    def coordinated_validate(session, locked_apartment_id, period, **kwargs):
        original_validate(session, locked_apartment_id, period, **kwargs)
        if period == date(2024, 3, 1):
            draft_validated.set()
            assert import_attempted.wait(timeout=2)
            import_validated.wait(timeout=0.25)
        else:
            import_validated.set()

    monkeypatch.setattr(
        billing_service,
        "validate_invoice_chronology",
        coordinated_validate,
    )
    monkeypatch.setattr(
        importer_service,
        "validate_invoice_chronology",
        coordinated_validate,
    )

    def create_earlier_draft() -> int:
        with session_factory() as session:
            apartment = session.get(Apartment, apartment_id)
            return create_draft(
                session,
                apartment,
                date(2024, 3, 1),
                Decimal("44.000000"),
            ).id

    def import_later_history() -> str:
        assert draft_validated.wait(timeout=2)
        import_attempted.set()
        with session_factory() as session:
            apartment = session.get(Apartment, apartment_id)
            try:
                import_xlsx(session, apartment, FIXTURE.read_bytes())
            except ImportFormatError as error:
                session.rollback()
                return str(error)
        raise AssertionError("Concurrent import unexpectedly succeeded")

    with ThreadPoolExecutor(max_workers=2) as executor:
        draft_future = executor.submit(create_earlier_draft)
        assert draft_validated.wait(timeout=2)
        import_future = executor.submit(import_later_history)
        assert draft_future.result(timeout=5) > 0
        import_error = import_future.result(timeout=5)

    assert "після незавершеної ранньої чернетки" in import_error
    assert not import_validated.is_set()
    with session_factory() as session:
        assert session.scalar(select(func.count(Invoice.id))) == 1
        assert session.scalar(select(func.count(Service.id))) == 0


async def test_import_rejects_historical_month_before_existing_invoice(tmp_path) -> None:
    application, lifespan, client = await _client(tmp_path)
    try:
        apartment_id = _apartment(application)
        engine = create_database_engine(application.state.settings.database_path)
        with create_session_factory(engine)() as session:
            apartment = session.get(Apartment, apartment_id)
            session.add(
                Invoice(
                    apartment=apartment,
                    period=date(2024, 6, 1),
                    status="paid",
                    exchange_rate=Decimal("44.000000"),
                    rent_amount_usd=Decimal("325.00"),
                    rent_amount_uah=Decimal("14300.00"),
                    utilities_total=Decimal("0.00"),
                    grand_total=Decimal("14300.00"),
                )
            )
            session.commit()
        engine.dispose()

        response = await _upload(client, apartment_id, FIXTURE.read_bytes())

        assert response.status_code == 422
        assert "перед наявним пізнішим рахунком" in response.json()["detail"]
        engine = create_database_engine(application.state.settings.database_path)
        with create_session_factory(engine)() as session:
            assert session.scalar(select(func.count(Invoice.id))) == 1
            assert session.scalar(select(func.count(Service.id))) == 0
        engine.dispose()
    finally:
        await _close(lifespan, client)


async def test_import_rejects_month_after_existing_draft(tmp_path) -> None:
    application, lifespan, client = await _client(tmp_path)
    try:
        apartment_id = _apartment(application)
        engine = create_database_engine(application.state.settings.database_path)
        with create_session_factory(engine)() as session:
            apartment = session.get(Apartment, apartment_id)
            session.add(
                Invoice(
                    apartment=apartment,
                    period=date(2024, 3, 1),
                    status="draft",
                    exchange_rate=Decimal("44.000000"),
                    rent_amount_usd=Decimal("325.00"),
                    rent_amount_uah=Decimal("14300.00"),
                    utilities_total=Decimal("0.00"),
                    grand_total=Decimal("14300.00"),
                )
            )
            session.commit()
        engine.dispose()

        response = await _upload(client, apartment_id, FIXTURE.read_bytes())

        assert response.status_code == 422
        assert "після незавершеної ранньої чернетки" in response.json()["detail"]
        engine = create_database_engine(application.state.settings.database_path)
        with create_session_factory(engine)() as session:
            assert session.scalar(select(func.count(Invoice.id))) == 1
            assert session.scalar(select(func.count(Service.id))) == 0
        engine.dispose()
    finally:
        await _close(lifespan, client)
