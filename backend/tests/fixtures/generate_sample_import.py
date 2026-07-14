"""Build the anonymized XLSX compatibility fixture.

The source export was unavailable in the repository. This generator preserves only
the layout artifacts the importer must support and contains no tenant data.
"""

from datetime import date
from pathlib import Path

from openpyxl import Workbook


def build_fixture(path: Path) -> None:
    workbook = Workbook()
    information = workbook.active
    information.title = "Загальна інформація"
    information.merge_cells("A1:F1")
    information["A1"] = "Анонімізована структура експорту"
    information.append([])
    information.append(
        ["Послуга", "Тип", "Одиниця", "Особовий рахунок", date(2024, 4, 1), date(2024, 5, 1)]
    )
    information.append(["Газ", "За лічильником", "м³", "ACC-XXXX", "7.95689грн.", "7,95689 грн"])
    information.append(["Утримання будинку", "Фіксована", "грн", "-", "210.51 грн.", "210.51 грн."])

    for title, current, fixed_amount in (
        ("Кві 2024", 122, "210.51 грн."),
        ("Тра 2024", 140, "не число"),
    ):
        sheet = workbook.create_sheet(title)
        sheet.merge_cells("A1:E1")
        sheet["A1"] = f"Розрахунок за {title}"
        sheet["A3"] = "Курс НБУ"
        sheet["B3"] = "44,68 грн."
        sheet["A4"] = "Оренда USD"
        sheet["B4"] = 325
        sheet["A5"] = "Оренда грн"
        sheet["B5"] = 14521
        sheet.append([])
        sheet.append(["Послуга", "Попередній", "Поточний", "Тариф", "Сума"])
        previous = 100 if title.startswith("Кві") else 122
        sheet.append(["Газ", previous, current, "7.95689грн.", (current - previous) * 7.95689])
        sheet.append(["Утримання будинку", "-", "-", "-", fixed_amount])

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


if __name__ == "__main__":
    build_fixture(Path(__file__).with_name("sample_import.xlsx"))
